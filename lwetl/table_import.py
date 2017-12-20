"""
    Table importer classes to extract dictionaries from eithr a CSV file, or and xls worksheet
"""

import base64
import binascii
import os
import re
import sys

from io import TextIOWrapper, StringIO
from codecs import EncodedFile

from csv import reader
from openpyxl import load_workbook

from .exceptions import EmptyFileError
from .utils import is_empty

RE_START_WITH_CHAR = re.compile('^[A-Z_].*')
LDAP_ATTR_MATCH = re.compile(r'^(?P<attr>[A-Za-z]+[0-9A-Za-z\-]*)(?P<sep>:{1,2})(?P<val>.*)')


def get_xls_row_values(row):
    return [cell.value for cell in row]


def unique_column_name(name, defined_columns: list) -> str:
    """
    Makes sure column names are unique in upper case
    @param name: str|None - suggested name
    @param defined_columns: list of defined columns
    @return: str - modified suggested names (upper case + counter, if required)
    @rtype: str
    """
    if is_empty(name):
        name = 'C%d' % (len(defined_columns) + 1)
    else:
        if not isinstance(name, str):
            name = str(name)
        name = name.strip().upper()
        if RE_START_WITH_CHAR.match(name) is None:
            name = 'C%d' % (len(defined_columns) + 1)

    x = 0
    v = name
    while v in defined_columns:
        x += 1
        v = '%s%d' % (name, x)
    return name


class BaseTextImport:

    def __init__(self, filename_or_stream=None, encoding='utf-8'):
        self.filename_or_stream = filename_or_stream
        self.encoding = self._check_encoding(encoding)

        self.fname = None
        self.fstream = None

    @staticmethod
    def _check_encoding(encoding):
        if isinstance(encoding,str) and (len(encoding.strip()) > 0):
            return encoding.strip()
        else:
            return 'utf-8'

    @staticmethod
    def _open_text_input(filename_or_stream, encoding='utf-8'):
        """
        Parse the input and determine if its a stream or a file
        @param filename_or_stream:
        @return: tuple one of which is
        """

        if is_empty(filename_or_stream):
            fstream = sys.stdin
            fname = None
        elif isinstance(filename_or_stream, str):
            fname = filename_or_stream
            if fname.lower() == 'stdin':
                fstream = sys.stdin
                fname = None
            else:
                fstream = open(fname, 'r', encoding=encoding)
        elif isinstance(filename_or_stream,(TextIOWrapper, StringIO, EncodedFile)):
            fname = None
            fstream = filename_or_stream
        else:
            raise TypeError('The input specifier must be a string (filename) or stream. Found: ' +
                            type(filename_or_stream).__name__)
        return fstream, fname

    def _force_close(self):
        """
        Forse closing of previous pending input
        """
        if isinstance(self.fstream, TextIOWrapper) and (self.fstream != sys.stdin) and \
                (not getattr(self.fstream, 'closed', False)) and (self.fname is not None):
            self.fstream.close()
            self.fstream = None
            self.fname = None

    def open(self, filename_or_stream=None, encoding=None):
        if filename_or_stream is None:
            filename_or_stream = self.filename_or_stream
        if self._check_encoding(encoding) != 'utf-8':
            self.encoding = encoding
        self._force_close()
        self.fstream, self.fname = self._open_text_input(filename_or_stream, self.encoding)
        return self

    def close(self):
        if self.fname is not None:
            self.fstream.close()
            self.fstream = None
            self.fname = None


class CsvImport(BaseTextImport):
    """
    Open a CSV file and extract data by row in the form of a dictionary
    Expects the first row if the dictionary to contain the column names
    """

    def __init__(self, filename_or_stream=None, delimiter: str = "\t", encoding='utf-8'):
        super(CsvImport, self).__init__(filename_or_stream, encoding)

        self.delimiter = self._check_delimiter(delimiter)

        self.csv = None
        self.columns = None
        self.n_columns = 0
        self.row_count = 0

    @staticmethod
    def _check_delimiter(delimiter):
        if is_empty(delimiter):
            delimiter = "\t"
        elif not isinstance(delimiter, str):
            raise ValueError('The delimiter specifier must be a string. Found: ' + type(delimiter).__name__)
        return delimiter

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def open(self, filename_or_stream=None, delimiter: str = None, encoding=None):
        super(CsvImport, self).open(filename_or_stream, encoding)

        if delimiter is not None:
            delimiter = self._check_delimiter(delimiter)
        else:
            delimiter = self.delimiter
        self.csv = reader(self.fstream, delimiter=delimiter)

        self.row_count = 0
        for row in self.csv:
            self.row_count += 1
            self.columns = []
            for value in row:
                self.columns.append(unique_column_name(value, self.columns))
            break

        if self.columns is None:
            if self.fname is None:
                input_spec = 'input stream'
            else:
                input_spec = self.fname
            raise EmptyFileError("No data found for: '%s'" % input_spec)
        else:
            self.n_columns = len(self.columns)
            self.row_count += 1

        return self

    def close(self):
        if self.csv is not None:
            self.csv = None
        self.columns = None
        self.n_columns = 0
        self.row_count = 0
        super(CsvImport, self).close()

    def get_data(self, max_rows=1000):
        if self.csv is None:
            raise ValueError('I/O operation on a closed CSV reader.')
        if (not isinstance(max_rows, int)) or (max_rows < 1):
            max_rows = 1000

        while True:
            rows = []
            for row in self.csv:
                rows.append(row)
                if len(rows) >= max_rows:
                    break

            for row in rows:
                self.row_count += 1
                dd = dict()
                indx = 0
                for value in row:
                    if indx < self.n_columns:
                        key = self.columns[indx]
                    elif not is_empty(value):
                        key = unique_column_name(None, self.columns)
                        self.columns.append(key)
                        self.n_columns += 1
                    indx += 1
                    if is_empty(value):
                        continue
                    dd[key] = value
                if len(dd) > 0:
                    yield dd
            if len(rows) < max_rows:
                break


class LdifImport(BaseTextImport):
    """
    Source: https://www.ibm.com/support/knowledgecenter/en/SSVJJU_6.2.0/com.ibm.IBMDS.doc_6.2/admin_gd34.htm

    LDIF is used to represent LDAP entries in text form. The basic form of an LDIF entry is:

    dn: <distinguished name>
    <attrtype> : <attrvalue>
    <attrtype> : <attrvalue>
    ...

    A line can be continued by starting the next line with a single space or tab character, for example:

          dn: cn=John E Doe, o=University of Higher
           Learning, c=US

    Multiple attribute values are specified on separate lines, for example:

          cn: John E Doe
          cn: John Doe

    If an <attrvalue> contains a non-US-ASCII character, or begins with a space or a colon ':', the <attrtype> is followed
    by a double colon and the value is encoded in base-64 notation.
    For example, the value " begins with a space" would be encoded like this:

          cn:: IGJlZ2lucyB3aXRoIGEgc3BhY2U=

    Multiple entries within the same LDIF file are separated by a blank line. Multiple blank lines are considered a
    logical end-of-file.
    LDIF example

    Here is an example of an LDIF file containing three entries.

          dn: cn=John E Doe, o=University of High
           er Learning, c=US
          cn: John E Doe
          cn: John Doe
          objectclass: person
          sn: Doe

          dn: cn=Bjorn L Doe, o=University of High
           er Learning, c=US
          cn: Bjorn L Doe
          cn: Bjorn Doe
          objectclass: person
          sn: Doe

          dn: cn=Jennifer K. Doe, o=University of High
           er Learning, c=US
          cn: Jennifer K. Doe
          cn: Jennifer Doe
          objectclass: person
          sn: Doe
          jpegPhoto:: /9j/4AAQSkZJRgABAAAAAQABAAD/2wBDABALD
           A4MChAODQ4SERATGCgaGBYWGDEjJR0oOjM9PDkzODdASFxOQ
           ERXRTc4UG1RV19iZ2hnPk1xeXBkeFxlZ2P/2wBDARESEhgVG
           ...

    The jpegPhoto in Jennifer Doe's entry is encoded using base-64. The textual attribute values can also be specified
    in base-64 format. However, if this is the case, the base-64 encoding must be in the code page of the wire format
    for the protocol (that is, for LDAP V2, the IA5 character set and for LDAP V3, the UTF-8 encoding).
    """

    def __init__(self, filename_or_stream = None, separator=None, encoding='utf-8'):
        super(LdifImport,self).__init__(filename_or_stream, encoding)
        self.separator = self._check_separator(separator)

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    @staticmethod
    def _check_separator(separator):
        if isinstance(separator,str) and (not is_empty(separator)):
            return separator
        else:
            return None

    def open(self,filename_or_stream=None, separator=None, encoding=None):
        super(LdifImport,self).open(filename_or_stream, encoding)
        if self._check_separator(separator) is not None:
            self.separator = separator
        return self

    def get_data(self):
        def init_record():
            return dict(), None, '', False

        def parse_attribute(attribute, value, is_base64, record):
            if attribute is not None:
                if (len(value) > 0) and is_base64:
                    try:
                        value = base64.standard_b64decode(value)
                    except binascii.Error as be:
                        print('ERROR: ' + be)
                        value = ''
                    else:
                        try:
                            svalue = value.decode(self.encoding)
                            value = svalue
                        except UnicodeError:
                            pass
                if len(value) > 0:
                    if attribute not in record:
                        record[attribute] = []
                    record[attribute].append(value)
            return None, '', False

        def finalize_record(record:dict)->dict:
            r = dict()
            for k,v in record.items():
                if len(v) == 1:
                    r[k] = v[0]
                elif len(v) > 1:
                    if isinstance(self.separator,str) and (not is_empty(self.separator)):
                        r[k] = self.separator.join([str(s) for s in v])
                    else:
                        r[k] = v
            return r

        if (self.fstream is None) or getattr(self.fstream,'closed',False):
            return []

        line_nr = 0
        record, attr, value, is_base64 = init_record()
        for line in self.fstream.readlines():
            line_nr += 1
            if len(line.strip()) == 0:
                if len(record) == 0:
                    break

                parse_attribute(attr, value, is_base64, record)
                r = finalize_record(record)
                if len(r) > 0:
                    yield r
                attr, value, is_base64 = None, '', False
                record = dict()
            elif (line[0] in [" ","\t"]) and (attr is not None):
                value += line.rstrip()[1:]
            else:
                m = LDAP_ATTR_MATCH.match(line)
                if m:
                    parse_attribute(attr, value, is_base64, record)
                    groups = m.groupdict()
                    attr = groups['attr']
                    if groups['val']:
                        value = groups['val'].strip()
                    else:
                        value = ''
                    is_base64 = (len(groups['sep']) > 1)
        if attr:
            parse_attribute(attr, value, is_base64, record)
        r = finalize_record(record)
        if len(r) > 0:
            yield r


class XlsxImport:
    """
    Open an xls worksheet and extract the data by row in the form of a dictionary
    Expects the first row of the worksheet to contain the column names
    """

    def __init__(self, file_name: str, sheet_name: str = None):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.work_book = None
        self.work_sheet = None
        self.columns = None
        self.n_columns = 0
        self.row_count = 0

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def open(self, file_name: str = None, sheet_name: str = None):
        if file_name is None:
            file_name = self.file_name
        if file_name is None:
            raise ValueError('Input filename not specified')
        elif not isinstance(file_name, str):
            raise TypeError('Filename specifier of wrong type: ' + type(file_name).__name__)
        if not os.path.isfile(file_name):
            raise FileNotFoundError("No such file: '%s'" % file_name)

        self.work_book = load_workbook(file_name, read_only=True, data_only=True)

        sheet_names = self.work_book.get_sheet_names()
        if len(sheet_names) == 0:
            raise EmptyFileError("Workbook '%s' has no data." % file_name)

        self.work_sheet = None
        if sheet_name is None:
            sheet_name = self.sheet_name
        if is_empty(sheet_name):
            sheet_name = None

        if sheet_name is None:
            self.work_sheet = self.work_book.get_sheet_by_name(sheet_names[0])
        elif isinstance(sheet_name, str):
            valid_sheets = [n for n in sheet_names if n.upper() == sheet_name.upper()]
            if len(valid_sheets) > 0:
                self.work_sheet = self.work_book.get_sheet_by_name(valid_sheets[0])
        else:
            raise ValueError('Sheet name must be string. Found: ' + type(sheet_name).__name__)

        if self.work_sheet is None:
            raise LookupError("Cannot find work sheet: '%s'" % sheet_name)
        self.row_count = 0
        for row in self.work_sheet.rows:
            self.columns = []
            for v in get_xls_row_values(row):
                self.columns.append(unique_column_name(v, self.columns))
            self.n_columns = len(self.columns)
            self.row_count += 1
            break
        if self.columns is None:
            raise EmptyFileError("Worksheet '%s' has no data." % sheet_name)
        return self

    def close(self):
        self.work_book = None
        self.work_sheet = None
        self.columns = None
        self.n_columns = 0
        self.row_count = 0

    def get_data(self, max_rows=1000):
        if self.work_sheet is None:
            raise LookupError('No active worksheet for data.')
        if (not isinstance(max_rows, int)) or (max_rows < 1):
            max_rows = 1000

        while True:
            rows = []
            n = self.row_count + 1
            for row in self.work_sheet.iter_rows(min_row=n, max_row=n + max_rows):
                rows.append(row)
            for row in rows:
                self.row_count += 1

                dd = dict()
                indx = 0
                for value in get_xls_row_values(row):
                    if indx < self.n_columns:
                        key = self.columns[indx]
                    else:
                        key = 'C%d' % indx
                    indx += 1
                    if is_empty(value):
                        continue
                    dd[key] = value
                if len(dd) > 0:
                    yield dd
            if len(rows) < max_rows:
                break
