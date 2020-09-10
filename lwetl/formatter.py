"""
    Bundles output formatters
"""

import csv
import sys

from collections import OrderedDict
from decimal import Decimal

# noinspection PyUnresolvedReferences
from xml.dom import minidom
from xml.etree.ElementTree import Element, SubElement, tostring

from jaydebeapi import Cursor

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.styles.fills import PatternFill

from .jdbc import Jdbc, DummyJdbc, PARENT_CONNECTION, COLUMN_TYPE_DATE, get_columns_of_cursor
from .uploader import NativeUploader, UPLOAD_MODE_PIPE
from .utils import *


def parse_output_selector(filename_or_stream):
    if is_empty(filename_or_stream):
        return sys.stdout, None
    elif isinstance(filename_or_stream, str):
        if filename_or_stream.lower() == 'stdout':
            return sys.stdout, None
        elif filename_or_stream == 'stderr':
            return sys.stderr, None
        else:
            return None, filename_or_stream
    elif type(filename_or_stream).__name__ in ['TextIOWrapper', 'StringIO', 'EncodedFile']:
        return filename_or_stream, None
    else:
        raise ValueError('Illegal type for the output specifier: ' + type(filename_or_stream).__name__)


class Formatter:
    def __init__(self, *args, **kwargs):
        self.cursor = kwargs.get('cursor', None)
        self.filename_or_stream = kwargs.get('filename_or_stream', sys.stdout)
        self.append = kwargs.get('append', False)

        self.fname = None
        self.fstream = None
        self.columns = None
        self.n_columns = 0

    def __enter__(self):
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def _call(self, *args, **kwargs):
        jdbc = kwargs.get('jdbc',None)
        if not isinstance(jdbc, Jdbc):
            raise ValueError('Jdbc connection of wrong type. Expected Jdbc, found: ' + type(jdbc).__name__)

        sql = kwargs.get('sql',None)
        if is_empty(sql) or (not isinstance(sql,str)):
            raise ValueError('SQL empty or not a string: ' + type(sql).__name__)
        self.cursor = jdbc.execute(sql,cursor=None)

        new_kwargs = {}
        for k,v in kwargs.items():
            if k not in ['cursor', 'sql', 'jdbc']:
                new_kwargs[k] = v
        self.open(*args, **new_kwargs)
        self.header()
        for row in jdbc.get_data(self.cursor):
            self.write(row)
        self.footer()
        self.close()

    def open(self, *args, **kwargs):
        cursor = kwargs.get('cursor', self.cursor)
        if not isinstance(cursor, Cursor):
            raise ValueError('Illegal cursor specifier: ' + type(cursor).__name__)

        self.columns = get_columns_of_cursor(cursor)
        self.n_columns = len(self.columns)

        append = verified_boolean(kwargs.get('append', self.append))

        filename_or_stream = kwargs.get('filename_or_stream', self.filename_or_stream)
        if filename_or_stream is None:
            outp = self.fstream if self.fname is None else self.fname
            self.fstream, self.fname = parse_output_selector(outp)
        else:
            self.fstream, self.fname = parse_output_selector(filename_or_stream)

        if is_empty(self.fname):
            self.fname = None
            if self.fstream is None:
                self.fstream = sys.stdout
        else:
            mode = 'a' if append else 'w'
            self.fstream = open(self.fname, mode)

        if type(self.fstream).__name__ not in ['TextIOWrapper', 'StringIO', 'EncodedFile']:
            raise IOError('Failed to open a output stream for formatter: ' + type(self.fstream).__name__)
        return self

    def close(self):
        if (self.fname is not None) and (self.fstream is not None) and (self.fstream not in [sys.stdout, sys.stderr]):
            self.fstream.close()
            self.fstream = None
        self.columns = None
        self.n_columns = None

    def format(self, row):
        if (self.n_columns is None) and (self.columns is not None):
            self.n_columns = len(self.columns)
        if (row is None) or (not isinstance(row, (list, tuple))) or (len(row) == 0):
            return []
        elif len(row) != self.n_columns:
            raise ValueError(
                'Mismatch in the number of columns: expected = %d, found = %d.' % (self.n_columns, len(row)))
        return [('Binary data suppressed.' if isinstance(r, bytes) else r) for r in row]

    def header(self):
        if self.fstream is not None:
            print(list(self.columns.keys()), file=self.fstream)

    def write(self, row: list):
        if self.fstream is not None:
            formatted_row = self.format(row)
            if formatted_row is not None:
                print(formatted_row, file=self.fstream)

    def footer(self):
        pass


class TextFormatter(Formatter):
    def __init__(self, *args, **kwargs):
        super(TextFormatter, self).__init__(*args, **kwargs)
        self.column_width, self.format_left, self.format_right = self._set_column_width(kwargs.get('column_width', 20))

    def __call__(self, *args, **kwargs):
        self._call(*args, **kwargs)

    def __enter__(self):
        return super(TextFormatter, self).__enter__()

    def __exit__(self, exc_type, exc_val, exc_tb):
        super(TextFormatter, self).__exit__(exc_type, exc_val, exc_tb)

    def open(self, *args, **kwargs):
        super(TextFormatter, self).open(*args, **kwargs)
        if 'column_width' in kwargs:
            self.column_width, self.format_left, self.format_right = self._set_column_width(kwargs['column_width'])
        return self

    @staticmethod
    def _set_column_width(column_width):
        if not isinstance(column_width, int):
            raise ValueError('Column width must be defined as an integer.')
        if column_width < 5:
            column_width = 5
        return column_width, '%%-%ds' % column_width, '%%%ds' % column_width

    def _limit_str(self, value, ctype):
        if (value is None) or (isinstance(value, str) and (len(value.strip()) == 0)):
            return self.format_left % ''
        elif isinstance(value, int):
            return self.format_right % str(value)
        elif isinstance(value, Decimal):
            return self.format_right % str(value)
        elif isinstance(value, str):
            v = value
        else:
            v = str(value)
        if len(v) > self.column_width:
            return v[:self.column_width - 3] + '...'
        else:
            return self.format_left % v

    def header(self):
        if self.fstream is not None:
            print(' | '.join([self._limit_str(c, 'str') for c in self.columns.keys()]), file=self.fstream)

    def format(self, row: list):
        row = super(TextFormatter, self).format(row)

        r = []
        x = 0
        for column_type in self.columns.values():
            r.append(self._limit_str(row[x], column_type))
            x += 1
        return ' | '.join(r)


class CsvFormatter(Formatter):
    def __init__(self, *args, **kwargs):
        super(CsvFormatter, self).__init__(*args, **kwargs)
        self.delimiter = kwargs.get('separator', ";")
        self.writer = None

    def __call__(self, *args, **kwargs):
        self._call(*args, **kwargs)

    def __enter__(self):
        return super(CsvFormatter, self).__enter__()

    def __exit__(self, exc_type, exc_val, exc_tb):
        super(CsvFormatter, self).__exit__(exc_type, exc_val, exc_tb)

    def open(self, *args, **kwargs):
        super(CsvFormatter, self).open(*args, **kwargs)
        self.delimiter = kwargs.get('separator', self.delimiter)
        self.writer = csv.writer(self.fstream, dialect='excel', delimiter=self.delimiter)
        return self

    def close(self):
        self.writer = None
        super(CsvFormatter, self).close()

    def header(self):
        if self.writer is None:
            raise ValueError('I/O operation not permitted on a closed CsvFormatter.')
        self.writer.writerow(list(self.columns.keys()))

    def write(self, row: list):
        if self.writer is None:
            raise ValueError('I/O operation not permitted on a closed CsvFormatter.')
        self.writer.writerow(self.format(row))


class XmlFormatter(Formatter):
    # supported formats
    PLAIN = 'plain'
    VALUE = 'value'
    EXCEL = 'excel'

    def __init__(self, *args, **kwargs):
        super(XmlFormatter, self).__init__(*args, **kwargs)
        self.name = kwargs.get('sheet_name', "Sheet1")
        self.dialect = kwargs.get('dialect', self.EXCEL)
        self.pretty_print = kwargs.get('pretty_print', False)

        self.book = None
        self.sheet = None
        self.table = None
        self.count = 0

    def __call__(self, *args, **kwargs):
        self._call(*args, **kwargs)

    def __enter__(self):
        return super(XmlFormatter, self).__enter__()

    def __exit__(self, exc_type, exc_val, exc_tb):
        super(XmlFormatter, self).__exit__(exc_type, exc_val, exc_tb)

    def set_dialect(self, dialect):
        if dialect not in [self.PLAIN, self.VALUE, self.EXCEL]:
            self.dialect = self.PLAIN
        else:
            self.dialect = dialect
        return self.dialect

    def open(self, *args, **kwargs):
        super(XmlFormatter, self).open(*args, **kwargs)
        sheet_name = kwargs.get('sheet_name', self.name)
        self.dialect = kwargs.get('dialect', self.dialect)
        self.pretty_print = verified_boolean(kwargs.get('pretty_print', self.pretty_print))

        self.count = 0
        if self.dialect == self.EXCEL:
            self.book = Element("Workbook", attrib={
                'xmlns': "urn:schemas-microsoft-com:office:spreadsheet",
                'xmlns:o': "urn:schemas-microsoft-com:office:office",
                'xmlns:x': "urn:schemas-microsoft-com:office:excel",
                'xmlns:ss': "urn:schemas-microsoft-com:office:spreadsheet",
                'xmlns:html': "http://www.w3.org/TR/REC-html40"})
            SubElement(self.book, "DocumentProperties", xmlns="urn:schemas-microsoft-com:office:office")
        else:
            self.book = Element("Document")

        if not is_empty(sheet_name):
            sheet_name = str(sheet_name)
        self.next_sheet(self.cursor,sheet_name)
        return self

    def close(self, pretty_print=None):
        def prettify(book):
            """
            Return a pretty-printed XML string for the Element.
            """
            rough_string = tostring(book, 'utf-8')
            reparsed = minidom.parseString(rough_string)
            return reparsed.toprettyxml(indent="  ")

        if self.book is None:
            raise ValueError('Cannot close a closed XmlFormatter.')

        if pretty_print is None:
            pretty_print = self.pretty_print
        else:
            pretty_print = verified_boolean(pretty_print)
        if self.fstream is not None:
            if pretty_print:
                print(prettify(self.book), file=self.fstream)
            else:
                print(tostring(self.book), file=self.fstream)
        self.book = None
        self.sheet = None
        super(XmlFormatter, self).close()

    def next_sheet(self, new_cursor=None, sheet_name=None):
        self.count += 1
        if isinstance(sheet_name, str) and (not is_empty(sheet_name)):
            self.name = sheet_name
        else:
            self.name = "Sheet%d" % self.count
        if self.dialect == self.EXCEL:
            self.sheet = SubElement(self.book, "ss:Worksheet", attrib={'ss:Name': self.name})
            self.table = SubElement(self.sheet, "Table")
        else:
            self.table = SubElement(self.book, "Table")

        if new_cursor is not None:
            if not isinstance(new_cursor, Cursor):
                raise ValueError('Cursor in next_sheet has wrong type: ' + type(new_cursor).__name__)
            self.cursor = new_cursor
            self.columns = get_columns_of_cursor(new_cursor)
            self.n_columns = len(self.columns)

    def header(self):
        if self.table is None:
            raise ValueError('I/O operation on closed xml table.')
        xml_row = SubElement(self.table, "Row")
        for column_name in self.columns.keys():
            if self.dialect == self.PLAIN:
                SubElement(xml_row, column_name).text = column_name
            elif self.dialect == self.VALUE:
                SubElement(xml_row, column_name, value=column_name)
            elif self.dialect == self.EXCEL:
                cell = SubElement(xml_row, "Cell")
                SubElement(cell, "Data").text = column_name
            else:
                raise KeyError('Illegal XLM dialect: ' + str(self.dialect))

    def write(self, row: list):
        if self.table is None:
            raise ValueError('I/O operation on closed xml table.')

        r = super(XmlFormatter, self).format(row)
        d = OrderedDict()
        for c in self.columns.keys():
            value = r.pop(0)
            if value is None:
                value = ''
            elif not isinstance(value, str):
                value = str(value)
            d[c] = value

        xml_row = SubElement(self.table, "Row")
        if self.dialect == self.PLAIN:
            for name, value in d.items():
                cell = SubElement(xml_row, name)
                if len(value.strip()) > 0:
                    cell.text = value
        elif self.dialect == self.VALUE:
            for name, value in d.items():
                SubElement(xml_row, name, value=value)
        elif self.dialect == self.EXCEL:
            for value in d.values():
                cell = SubElement(xml_row, "Cell")
                if len(value.strip()) > 0:
                    SubElement(cell, "Data").text = value
        else:
            raise KeyError('Illegal XLM dialect: ' + str(self.dialect))


class XlsxFormatter(Formatter):
    def __init__(self, *args, **kwargs):
        super(XlsxFormatter, self).__init__(*args, **kwargs)
        self.name = kwargs.get('sheet_name', "Sheet1")
        self.pretty = kwargs.get('pretty', False)
        self.book = None
        self.sheet = None
        self.count = 0

    def __call__(self, *args, **kwargs):
        self._call(*args, **kwargs)

    def __enter__(self):
        return super(XlsxFormatter, self).__enter__()

    def __exit__(self, exc_type, exc_val, exc_tb):
        super(XlsxFormatter, self).__exit__(exc_type, exc_val, exc_tb)

    def open(self, *args, **kwargs):
        super(XlsxFormatter, self).open(*args, **kwargs)
        sheet_name = kwargs.get('sheet_name', self.name)
        if self.fname is None:
            raise AttributeError('XLS files can only be written to file.')
        else:
            self.fstream.close()
            self.fstream = None

        self.count = 0
        self.book = Workbook(write_only=True)
        self.sheets_with_header = set()
        self.next_sheet(self.cursor,sheet_name)

        return self

    def close(self):
        if self.book is None:
            raise ValueError('Cannot close a closed XlsxFormatter.')
        self.book.save(self.fname)
        self.book = None
        self.sheet = None
        if self.pretty:
            self.prettify()
        super(XlsxFormatter, self).close()

    def next_sheet(self, new_cursor=None, sheet_name=None):
        if self.count > 0:
            self.name = None
        self.count += 1

        if (isinstance(sheet_name, str) and (len(sheet_name.strip()) > 0)):
            self.name = sheet_name
        if not (isinstance(self.name, str) and (len(self.name.strip()) > 0)):
            self.name = "Sheet%d" % self.count

        self.sheet = self.book.create_sheet(title=self.name)

        if new_cursor is not None:
            if not isinstance(new_cursor, Cursor):
                raise ValueError('Cursor in next_sheet has wrong type: ' + type(new_cursor).__name__)
            self.cursor = new_cursor
            self.cursor = new_cursor
            self.columns = get_columns_of_cursor(new_cursor)
            self.n_columns = len(self.columns)

    def header(self):
        if self.sheet is None:
            raise ValueError('I/O operation on closed xlsx formatter.')
        self.sheets_with_header.add(str(self.sheet.title))
        self.sheet.append(list(self.columns.keys()))

    def write(self, row: list):
        if self.sheet is None:
            raise ValueError('I/O operation on closed xlsx formatter.')
        r = super(XlsxFormatter, self).format(row)
        x = 0
        for column_type in self.columns.values():
            if isinstance(r[x], str) and (column_type == COLUMN_TYPE_DATE):
                r[x] = string2date(r[x])
            x += 1
        self.sheet.append(r)

    def prettify(self):
        b = load_workbook(self.fname)
        header_style = NamedStyle(name='table_header')
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(fill_type='solid', start_color='00CCCCCC', end_color='00CCCCCC')
        header_style.alignment = Alignment(horizontal='center')
        thin = Side(border_style='thin', color='000000')
        double = Side(border_style="double", color="ff0000")
        header_style.border = Border(top=double, left=thin, right=thin, bottom=double)

        b.add_named_style(header_style)

        for ws in b.worksheets:
            if ws.title in self.sheets_with_header:
                has_header = True
                for cell in ws[1]:
                    cell.style = header_style
            else:
                has_header = False
            for x, c in enumerate(ws.columns, start=1):
                column_width = 5
                for y, cell in enumerate(c, start=1):
                    if cell.value is not None:
                        w = len(str(cell.value))
                        if has_header and (y == 1):
                            w = int(1.25*w)
                        if w > column_width:
                            column_width = w
                if column_width > 50:
                    column_width = 50
                ws.column_dimensions[get_column_letter(x)].width = column_width+2
        b.save(self.fname)

class SqlFormatter(Formatter):
    UNDEFINED_TABLE = '@UNDEFINED_TABLE@'

    def __init__(self, *args, **kwargs):
        super(SqlFormatter, self).__init__(*args, **kwargs)
        self.jdbc = kwargs.get('connection', None)
        self.table_name = kwargs.get('table', self.UNDEFINED_TABLE)
        self.columns = kwargs.get('columns', None)
        self.database_type = kwargs.get('type', None)

        self.uploader = None

    def __call__(self, *args, **kwargs):
        self._call(*args, **kwargs)

    def __enter__(self):
        outp = self.fstream if self.fname is None else self.fname
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb):
        super(SqlFormatter, self).__exit__(exc_type, exc_val, exc_tb)

    def open(self, *args, **kwargs):
        target_columns = kwargs.get('columns', self.columns)
        db_type = kwargs.get('type',self.database_type)
        super(SqlFormatter, self).open(*args, **kwargs)

        connection = kwargs.get('connection', self.jdbc)
        if not isinstance(connection, Jdbc):
            database_type = kwargs.get('type', self.database_type)
            if database_type is None:
                connection = getattr(self.cursor, PARENT_CONNECTION)
            else:
                connection = DummyJdbc(database_type)

        if not isinstance(connection, (Jdbc, DummyJdbc)):
            raise RuntimeError('Target database for upload cannot be defined. Please check your input parameters.')

        t_name = kwargs.get('table', self.table_name)
        if db_type is None:
            db_type = connection.type

        self.uploader = NativeUploader(connection, t_name, fstream=None, commit_mode=UPLOAD_MODE_PIPE,
                                       columns=target_columns, type=db_type)
        return self

    def close(self):
        self.uploader = None
        super(SqlFormatter, self).close()

    def header(self):
        pass

    def format(self, row: list):
        if self.uploader is None:
            raise ValueError('I/O operation on closed sql formatter.')
        x = 0
        dd = dict()
        rs = super(SqlFormatter, self).format(row)
        for column_name, type in self.columns.items():
            value = rs[x]
            if not is_empty(value):
                dd[column_name] = value
            x += 1
        self.uploader.insert(dd)
        sqls = self.uploader.commit()
        if self.fstream is not None:
            print(';\n'.join(sqls) + ';', file=self.fstream)
